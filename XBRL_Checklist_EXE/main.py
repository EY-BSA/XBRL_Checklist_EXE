"""
XBRL CoE Checklist (독립 실행형)
IxD 편집기 '구조내려받기' xlsx → 전체 검증 (1-1 ~ 7-2) → XBRL_CoE_Checklist_Result.xlsx 저장
"""

import io
import re
import sys
import os
import threading
import tkinter as tk
import zipfile
from tkinter import filedialog, messagebox, ttk

import openpyxl

from taxonomy_xlsx_parser import parse_taxonomy_xlsx
from checklist_engine import run_all_checks, CheckResult
from standard_taxonomy import StandardTaxonomy, enrich_axis_domain_check


def resource_path(relative: str) -> str:
    """PyInstaller 번들 실행 시 내장 리소스 경로 반환."""
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative)


AXIS_DOMAIN_PATH = resource_path('Axis_Domain_Check.xlsx')
TEMPLATE_PATH    = resource_path(os.path.join('template', 'XBRL_CoE_Checklist_Result.xlsx'))


# ─── 엑셀 출력 헬퍼 ──────────────────────────────────────────────────────────

def _role_def(iss) -> str:
    rc = iss.role_code; ko = iss.role_name_ko; en = iss.role_name_en
    if rc and ko:
        return f"[{rc}] {ko} | {en}" if en else f"[{rc}] {ko}"
    return ko or rc


def _table_name(iss) -> str:
    return iss.table_name_ko or iss.role_name_ko or iss.role_code


def _find_sheet_xml(xlsx_path: str, sheet_name: str) -> str:
    """템플릿 xlsx에서 시트 이름에 해당하는 worksheet XML 경로 반환."""
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        wb_xml   = z.read('xl/workbook.xml').decode('utf-8')
        rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')

    escaped = re.escape(sheet_name)
    m = re.search(rf'<sheet\b[^>]*\bname="{escaped}"[^>]*/>', wb_xml)
    if not m:
        raise KeyError(f'workbook.xml에서 시트 찾기 실패: {sheet_name}')
    rid_m = re.search(r'r:id="([^"]+)"', m.group(0))
    if not rid_m:
        raise KeyError(f'<sheet> 태그에서 r:id 추출 실패: {sheet_name}')
    rid = rid_m.group(1)

    m2 = re.search(rf'<Relationship\b[^>]*\bId="{re.escape(rid)}"[^>]*/>', rels_xml)
    if not m2:
        raise KeyError(f'workbook.xml.rels에서 rId 찾기 실패: {rid}')
    target_m = re.search(r'Target="worksheets/([^"]+)"', m2.group(0))
    if not target_m:
        raise KeyError(f'<Relationship> 태그에서 Target 추출 실패: {rid}')
    return f'xl/worksheets/{target_m.group(1)}'


_DRAWING_PAT = re.compile(rb'<(?:legacyDrawing|drawing)[^>]*/>')


def _write_standard_row(ws, ri: int, iss):
    ws.cell(ri, 2, _role_def(iss))
    ws.cell(ri, 3, _table_name(iss))
    ws.cell(ri, 4, iss.prefix)
    ws.cell(ri, 5, iss.element_name)
    ws.cell(ri, 6, iss.label_ko)
    ws.cell(ri, 7, iss.label_en)
    ws.cell(ri, 8, iss.label_role)
    ws.cell(ri, 9, iss.data_type)
    ws.cell(ri, 10, iss.period)


def _write_sheet(ws, result: CheckResult):
    """openpyxl worksheet 객체에 검증 결과를 기록한다."""
    if ws.max_row >= 14:
        for row in ws.iter_rows(min_row=14, max_row=ws.max_row):
            for cell in row:
                cell.value = None

    chk_id = result.check_id
    for ri, iss in enumerate(result.issues, 14):
        if chk_id == '5-6':
            ws.cell(ri, 2, _table_name(iss))
            ws.cell(ri, 3, '단위미표시')
        elif chk_id == '7-1':
            _write_standard_row(ws, ri, iss)
            ws.cell(ri, 11, iss.dart_negate)
            ws.cell(ri, 12, iss.client_negate)
        else:
            _write_standard_row(ws, ri, iss)

    ws['B11'] = result.issue_count


def write_all_results(results: dict, output_path: str):
    """모든 체크 결과를 템플릿 기반 xlsx에 기록한다."""
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f'템플릿을 찾을 수 없습니다: {TEMPLATE_PATH}')

    # Step 1: openpyxl로 모든 시트에 데이터 기록 후 메모리 버퍼에 저장
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    for check_id, result in results.items():
        sheet_name = result.sheet
        if sheet_name not in wb.sheetnames:
            continue
        _write_sheet(wb[sheet_name], result)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Step 2: 수정된 시트 XML 목록 수집 (drawing/legacyDrawing 태그 복원용)
    modified_xmls: dict[str, list] = {}  # xml_path → [drawing tag bytes, ...]
    with zipfile.ZipFile(TEMPLATE_PATH, 'r') as tmpl_zip:
        for check_id, result in results.items():
            try:
                xml_path = _find_sheet_xml(TEMPLATE_PATH, result.sheet)
            except KeyError:
                continue
            try:
                raw = tmpl_zip.read(xml_path)
                modified_xmls[xml_path] = _DRAWING_PAT.findall(raw)
            except KeyError:
                modified_xmls[xml_path] = []

    preserve_from_modified = set(modified_xmls.keys()) | {'xl/sharedStrings.xml', 'xl/styles.xml'}
    # 수정된 시트의 rels 파일도 openpyxl 버전 사용 (rId 재번호 부여 동기화)
    for xml_path in modified_xmls.keys():
        parts = xml_path.rsplit('/', 1)
        preserve_from_modified.add(f'{parts[0]}/_rels/{parts[1]}.rels')

    with zipfile.ZipFile(TEMPLATE_PATH, 'r') as tmpl_zip, \
         zipfile.ZipFile(buf, 'r') as mod_zip:
        mod_names = set(mod_zip.namelist())

        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as out_zip:
            for info in tmpl_zip.infolist():
                if info.filename == 'xl/calcChain.xml':
                    continue
                raw = tmpl_zip.read(info.filename)
                if info.filename in preserve_from_modified and info.filename in mod_names:
                    raw = mod_zip.read(info.filename)
                    # openpyxl이 drawing/legacyDrawing 태그를 제거한 경우에만 복원 (중복 방지)
                    if info.filename in modified_xmls:
                        drawing_tags = modified_xmls[info.filename]
                        if drawing_tags and not _DRAWING_PAT.search(raw):
                            inject = b''.join(drawing_tags)
                            ws_tag_end = raw.find(b'>', raw.find(b'<worksheet'))
                            if b'xmlns:r=' not in raw[:ws_tag_end]:
                                ns = b' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
                                inject = re.sub(
                                    rb'<((?:legacyDrawing|drawing)) ',
                                    rb'<\1' + ns + rb' ',
                                    inject
                                )
                            raw = raw.replace(b'</worksheet>', inject + b'</worksheet>')
                if info.filename == '[Content_Types].xml':
                    raw = re.sub(rb'<Override[^>]*calcChain[^>]*/>', b'', raw)
                if info.filename == 'xl/workbook.xml':
                    raw = re.sub(rb'<calcPr([^/]*)/>', rb'<calcPr\1 fullCalcOnLoad="1"/>', raw)
                out_zip.writestr(info, raw)


# ─── 분석 로직 ────────────────────────────────────────────────────────────────

def run_check(input_path: str):
    with open(input_path, 'rb') as f:
        xlsx_bytes = f.read()

    data = parse_taxonomy_xlsx(xlsx_bytes)

    std = StandardTaxonomy()
    if os.path.exists(AXIS_DOMAIN_PATH):
        enrich_axis_domain_check(std, AXIS_DOMAIN_PATH)

    results = run_all_checks(data, std)
    return data, results


# ─── GUI ──────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('XBRL CoE Checklist')
        self.resizable(False, False)
        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f'+{(sw-w)//2}+{(sh-h)//2}')

    def _build_ui(self):
        pad = {'padx': 16, 'pady': 8}

        tk.Label(self, text='XBRL CoE Checklist', font=('Helvetica', 15, 'bold')).pack(**pad)
        tk.Label(self, text='전체 검증  1-1 ~ 7-2  (29개 체크)', font=('Helvetica', 11)).pack(pady=(0, 4))
        tk.Label(
            self,
            text='IxD 편집기 [구조내려받기] xlsx 파일을 선택하세요.',
            fg='#555'
        ).pack()

        self._file_var = tk.StringVar(value='선택된 파일 없음')
        tk.Label(self, textvariable=self._file_var, fg='#333', wraplength=420).pack(**pad)

        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=4)
        tk.Button(btn_frame, text='📂  파일 선택', width=18,
                  command=self._select_file).grid(row=0, column=0, padx=6)
        self._run_btn = tk.Button(btn_frame, text='▶  체크 실행', width=18,
                                  state='disabled', command=self._run)
        self._run_btn.grid(row=0, column=1, padx=6)

        self._progress = ttk.Progressbar(self, mode='indeterminate', length=420)
        self._progress.pack(pady=(8, 2))

        self._status = tk.StringVar(value='')
        tk.Label(self, textvariable=self._status, fg='#333').pack(pady=(0, 12))

        self._input_path = None

    def _select_file(self):
        path = filedialog.askopenfilename(
            title='XBRL 구조내려받기 파일 선택',
            filetypes=[('Excel 파일', '*.xlsx'), ('모든 파일', '*.*')]
        )
        if not path:
            return
        self._input_path = path
        self._file_var.set(os.path.basename(path))
        self._run_btn.config(state='normal')
        self._status.set('')

    def _run(self):
        if not self._input_path:
            return
        self._run_btn.config(state='disabled')
        self._progress.start(12)
        self._status.set('분석 중...')
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        try:
            data, results = run_check(self._input_path)
            self.after(0, self._on_done, data, results)
        except Exception as exc:
            self.after(0, self._on_error, str(exc))

    def _on_done(self, data, results):
        self._progress.stop()

        total_issues = sum(r.issue_count for r in results.values())
        checks_with_issues = sum(1 for r in results.values() if not r.passed)

        company = data.company_name or 'XBRL'
        rdate   = (data.report_date or '').replace('/', '').replace('-', '')
        parts   = [p for p in [company, rdate] if p]
        default_name = f"XBRL_CoE_Checklist_Result_{'_'.join(parts)}.xlsx"

        if total_issues:
            msg = f'이슈 {total_issues}건 발견 ({checks_with_issues}개 항목).\n결과 파일을 저장할 위치를 선택하세요.'
        else:
            msg = '이슈 없음 (전체 Pass). 결과 파일을 저장할 위치를 선택하세요.'
        self._status.set(msg)

        output_path = filedialog.asksaveasfilename(
            title='결과 파일 저장',
            defaultextension='.xlsx',
            initialfile=default_name,
            filetypes=[('Excel 파일', '*.xlsx'), ('모든 파일', '*.*')]
        )
        if not output_path:
            self._run_btn.config(state='normal')
            self._status.set('저장 취소됨.')
            return

        try:
            write_all_results(results, output_path)
            self._status.set(f'저장 완료: {os.path.basename(output_path)}')

            # 시트별 요약 문자열
            summary_lines = []
            for cid, res in results.items():
                if not res.passed:
                    summary_lines.append(f'  {cid}: {res.issue_count}건')
            summary_str = '\n'.join(summary_lines) if summary_lines else '  (없음)'

            messagebox.showinfo(
                '완료',
                f'전체 검증 완료\n'
                f'총 이슈: {total_issues}건 / {checks_with_issues}개 항목\n\n'
                f'이슈 발생 항목:\n{summary_str}\n\n'
                f'저장 위치:\n{output_path}'
            )
        except Exception as exc:
            self._on_error(str(exc))
        finally:
            self._run_btn.config(state='normal')

    def _on_error(self, msg: str):
        self._progress.stop()
        self._status.set('오류 발생')
        self._run_btn.config(state='normal')
        messagebox.showerror('오류', msg)


if __name__ == '__main__':
    App().mainloop()
